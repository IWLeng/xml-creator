import openpyxl as op
import os

print("Welcome to XML Creator for Icario")

# Language selection dictionary
languages = {
    "1": {"name": "English", "tts": "azure", "voice": "en-US, AriaNeural"},
    "2": {"name": "Spanish", "tts": "amazon"},
    "3": {"name": "Chinese", "tts": "azure", "voice": "zh-CN, XiaochenMultilingualNeural"},
    "4": {"name": "Korean", "tts": "azure", "voice": "en-US, EmmaMultilingualNeural"},
    "5": {"name": "Vietnamese", "tts": "azure", "voice": "en-US, EmmaMultilingualNeural"},
    "6": {"name": "Armenian", "tts": "azure", "voice": "en-US, EmmaMultilingualNeural"}
}

# Azure template (fixed xml:lang="en-US")
azure_template = """<speak xmlns="http://www.w3.org/2001/10/synthesis" xmlns:mstts="http://www.w3.org/2001/mstts" xmlns:emo="http://www.w3.org/2009/10/emotionml" version="1.0" xml:lang="en-US"><voice name="Microsoft Server Speech Text to Speech Voice ({voice})"><prosody rate="-5%" pitch="0%"><mstts:silence type="Sentenceboundary" value="200ms"/>\n\n{content}\n<mstts:silence type="Tailing" value="200ms"/></prosody></voice></speak>"""

# Amazon template
amazon_template = """<speak><prosody rate="95%">\n\n{content}\n</prosody></speak>"""

def excel_process():
    folder_path = input("Please enter the folder path containing the .xlsx files: ")

    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

    for excel_file_name in excel_files:
        excel_file_path = os.path.join(folder_path, excel_file_name)
        wb = op.load_workbook(excel_file_path)
        base_dir = os.path.dirname(excel_file_path)

        print(f"Processing file: {excel_file_name}")

        # Display language options dynamically
        print("Select engine for this file:")
        for key, lang in languages.items():
            print(f"Press '{key}' for {lang['name']}")

        lang_choice = input("Enter your choice: ")

        if lang_choice not in languages:
            print("Invalid selection. Skipping this file.")
            continue

        lang_config = languages[lang_choice]
        excel_file_name_without_ext = os.path.splitext(excel_file_name)[0]

        # Check if there's only one sheet
        if len(wb.sheetnames) == 1:
            output_dir = os.path.join(base_dir, excel_file_name_without_ext)
        else:
            output_dir = os.path.join(base_dir, excel_file_name_without_ext)

        os.makedirs(output_dir, exist_ok=True)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Only create a subfolder if there are multiple sheets
            if len(wb.sheetnames) > 1:
                sheet_output_dir = os.path.join(output_dir, sheet_name)
            else:
                sheet_output_dir = output_dir

            os.makedirs(sheet_output_dir, exist_ok=True)

            print(f"Processing sheet: {sheet_name} into directory: {sheet_output_dir}")

            for row in sheet.iter_rows(min_row=1):
                filename = row[0].value
                content = row[1].value

                if filename and content:
                    filename = f"{filename}.xml"
                    filepath = os.path.join(sheet_output_dir, filename)

                    if lang_config["tts"] == "azure":
                        xml_content = azure_template.format(voice=lang_config["voice"], content=content)
                    else:
                        xml_content = amazon_template.format(content=content)

                    with open(filepath, 'w', encoding="utf-8") as file:
                        file.write(xml_content)

                    print(f"Created file: {filepath}")

    print("All files have been created.")

# Run the function
excel_process()
