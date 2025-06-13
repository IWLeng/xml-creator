import openpyxl as op
import os
import re  # Import re module for regex

# Function to sanitize file names by removing invalid characters
def sanitize_filename(filename):
    # Replace invalid characters with an underscore
    return re.sub(r'[\\/*?:"<>|]', '_', filename).replace('\n', '').replace('\r', '')

def process_excel_file(excel_file_path, engine_select, azure_voice):
    wb = op.load_workbook(excel_file_path)

    # Excel file dir
    base_dir = os.path.dirname(excel_file_path)
    excel_file_name = os.path.splitext(os.path.basename(excel_file_path))[0]

    azure_header1 = """<speak xmlns="http://www.w3.org/2001/10/synthesis" xmlns:mstts="http://www.w3.org/2001/mstts" xmlns:emo="http://www.w3.org/2009/10/emotionml" version="1.0" xml:lang="en-US"><voice name="Microsoft Server Speech Text to Speech Voice ("""
    azure_header2 = """)"><prosody rate="0%" pitch="0%">"""
    azure_footer = """</prosody></voice></speak>"""

    amazon_header = """<speak>"""
    amazon_footer = """</speak>"""

    # Process sheets
    if len(wb.sheetnames) == 1:
        # If there's only one sheet, use the Excel file name as the output directory
        output_dir = os.path.join(base_dir, excel_file_name)
        os.makedirs(output_dir, exist_ok=True)

        sheet = wb[wb.sheetnames[0]]
        print(f"Processing sheet: {sheet.title} into directory: {output_dir}")

        for row in sheet.iter_rows(min_row=1):
            filename = row[0].value
            content = row[1].value
            if filename:
                filename = sanitize_filename(f"{filename}.xml")  # Sanitize the filename
                filepath = os.path.join(output_dir, filename)

                if engine_select == 1:
                    combined = f"{azure_header1}{azure_voice}{azure_header2}\n\n{content}\n{azure_footer}"
                else:
                    combined = f"{amazon_header}\n\n{content}\n{amazon_footer}"

                if content:
                    with open(filepath, 'w', encoding="utf-8") as file:
                        file.write(combined)
                    print(f"Created file: {filepath}")
    else:
        # Process multiple sheets
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_output_dir = os.path.join(base_dir, excel_file_name, sheet_name)
            os.makedirs(sheet_output_dir, exist_ok=True)

            print(f"Processing sheet: {sheet_name} into directory: {sheet_output_dir}")

            for row in sheet.iter_rows(min_row=1):
                filename = row[0].value
                content = row[1].value
                if filename:
                    filename = sanitize_filename(f"{filename}.xml")  # Sanitize the filename
                    filepath = os.path.join(sheet_output_dir, filename)

                    if engine_select == 1:
                        combined = f"{azure_header1}{azure_voice}{azure_header2}\n\n{content}\n{azure_footer}"
                    else:
                        combined = f"{amazon_header}\n\n{content}\n{amazon_footer}"

                    if content:
                        with open(filepath, 'w', encoding="utf-8") as file:
                            file.write(combined)
                        print(f"Created file: {filepath}")

print("Welcome to XML Creator")

directory_path = input("Please enter the directory containing the Excel files: ")

# Engine selection
engine_select = int(input("Select engine. Press '1' for Azure; press '2' for Amazon/Google: "))
azure_voice = None
if engine_select == 1:
    azure_voice = input("Please input voice in format 'xx-XX, nameNeural' [for example: en-US, AriaNeural]: ")

# Process each .xlsx file in the directory
for file_name in os.listdir(directory_path):
    if file_name.endswith('.xlsx'):
        file_path = os.path.join(directory_path, file_name)
        print(f"Processing file: {file_path}")
        process_excel_file(file_path, engine_select, azure_voice)

print("All files have been created.")
input("Press Enter to exit.")
