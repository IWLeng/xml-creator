import openpyxl as op
import os
import re

def find_excel_files(root_folder):
    """Find all Excel files recursively in the given folder and its subfolders."""
    excel_files = []
    for root, dirs, files in os.walk(root_folder):
        for file in files:
            if file.endswith('.xlsx'):
                full_path = os.path.join(root, file)
                excel_files.append(full_path)
    return excel_files

def process_excel_file(excel_file_path, column_index, start_row, engine_select, azure_voice):
    """Process a single Excel file and create XML output."""
    try:
        wb = op.load_workbook(excel_file_path)
    except Exception as e:
        print(f"Error loading file {excel_file_path}: {e}")
        return None

    content_lines = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"  Processing sheet: {sheet_name}")

        for row in range(start_row, sheet.max_row + 1):
            cell_value = str(sheet.cell(row=row, column=column_index + 1).value or "").strip()
            if cell_value:
                # Add <break time="2000ms" /> after each line
                content_lines.append(f'{cell_value} <break time="2000ms" />')

    if not content_lines:
        print(f"  No content found. Skipping...")
        return None

    return content_lines

def process_folder(folder_path):
    """Process all Excel files in the folder and its subfolders."""
    print(f"\nSearching for Excel files in: {folder_path}")
    
    # Find all Excel files recursively
    excel_files = find_excel_files(folder_path)
    
    if not excel_files:
        print(f"No .xlsx files found in the folder or its subfolders: {folder_path}")
        return False

    print(f"Found {len(excel_files)} Excel file(s):")
    for file in excel_files:
        print(f"  - {os.path.relpath(file, folder_path)}")

    # Get column selection
    while True:
        column_input = input("\nEnter column letter (A-Z) or number (1-based) to extract text from: ").strip().upper()
        try:
            if column_input.isalpha():
                # Convert letter to column index (0-based)
                column_index = ord(column_input) - ord('A')
            else:
                # Convert number to column index (0-based)
                column_index = int(column_input) - 1
            
            if column_index < 0:
                raise ValueError("Column must be positive")
            break
        except (ValueError, TypeError) as e:
            print(f"Invalid column input: {e}. Please try again.")

    # Get start row
    while True:
        start_row_input = input("Enter starting row number (1-based, default is 2): ").strip()
        if not start_row_input:
            start_row = 2
            break
        try:
            start_row = int(start_row_input)
            if start_row < 1:
                raise ValueError("Row must be at least 1")
            break
        except ValueError as e:
            print(f"Invalid row input: {e}. Please try again.")

    # Engine selection
    while True:
        engine_select = input("Select engine. Press '1' for Azure; press '2' for Amazon/Google: ").strip()
        if engine_select in ['1', '2']:
            break
        print("Invalid engine selection. Please choose either '1' or '2'.")

    azure_voice = ""
    if engine_select == '1':
        while True:
            azure_voice = input("Please input voice in format 'xx-XX, nameNeural' [for example: en-US, AriaNeural]: ").strip()
            if azure_voice:
                break
            print("Voice input cannot be empty.")

    azure_header1 = """<speak xmlns="http://www.w3.org/2001/10/synthesis" xmlns:mstts="http://www.w3.org/2001/mstts" xmlns:emo="http://www.w3.org/2009/10/emotionml" version="1.0" xml:lang="en-US"><voice name="Microsoft Server Speech Text to Speech Voice ("""
    azure_header2 = """)"><prosody rate="0%" pitch="0%">"""
    azure_footer = """</prosody></voice></speak>"""

    amazon_header = """<speak>"""
    amazon_footer = """</speak>"""

    processed_count = 0
    skipped_count = 0
    
    print(f"\nProcessing {len(excel_files)} file(s)...")
    
    for excel_file_path in excel_files:
        print(f"\nProcessing: {os.path.relpath(excel_file_path, folder_path)}")
        
        # Get relative path for output
        relative_path = os.path.relpath(os.path.dirname(excel_file_path), folder_path)
        if relative_path == '.':
            output_dir = folder_path  # Root folder
        else:
            output_dir = os.path.join(folder_path, relative_path)
        
        # Process the Excel file
        content_lines = process_excel_file(excel_file_path, column_index, start_row, engine_select, azure_voice)
        
        if content_lines:
            content = "\n".join(content_lines)
            base_name = os.path.splitext(os.path.basename(excel_file_path))[0]
            filename = f"{base_name}.xml"
            filepath = os.path.join(output_dir, filename)

            if engine_select == '1':
                combined = f"{azure_header1}{azure_voice}{azure_header2}\n\n{content}\n{azure_footer}"
            else:
                combined = f"{amazon_header}\n\n{content}\n{amazon_footer}"

            with open(filepath, 'w', encoding="utf-8") as file:
                file.write(combined)

            print(f"  Created: {os.path.relpath(filepath, folder_path)}")
            processed_count += 1
        else:
            skipped_count += 1
    
    print(f"\nProcessing complete!")
    print(f"  Files processed: {processed_count}")
    print(f"  Files skipped: {skipped_count}")
    
    return True

def main():
    print("Welcome to XML Creator")
    print("This script creates XML files from Excel files with 2-second breaks between each row.")
    print("Note: This will process ALL Excel files in the specified folder AND all its subfolders.")

    while True:
        print("\n" + "="*50)
        folder_path = input("Please enter the folder path containing the .xlsx files: ").strip()
            
        # Validate folder path
        if not os.path.exists(folder_path):
            print("Error: The specified folder path does not exist.")
            continue
        if not os.path.isdir(folder_path):
            print("Error: The specified path is not a folder.")
            continue
            
        process_folder(folder_path)
        
        another = input("\nProcess another folder? (y/n): ").strip().lower()
        if another not in ['y', '1']:
            break

    print("\nThank you for using XML Creator!")
    input("Press Enter to exit.")

if __name__ == "__main__":
    main()