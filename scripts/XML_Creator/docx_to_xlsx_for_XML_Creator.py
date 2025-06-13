### UPDATE THE SCRIPT TO ITERATE ALL THE TABLES IN THE DOCUMENT


import os
import win32com.client
from docx import Document
import openpyxl

def convert_doc_to_docx(doc_path):
    """Convert .doc to .docx using Word COM API"""
    word = win32com.client.Dispatch("Word.Application")
    word.Visible = False  # Run in background
    
    docx_path = doc_path + "x"  # Append 'x' to change ".doc" -> ".docx"
    doc = word.Documents.Open(doc_path)
    doc.SaveAs(docx_path, FileFormat=16)  # 16 = wdFormatDocumentDefault (.docx)
    doc.Close()
    word.Quit()
    
    return docx_path

def process_docx_to_excel():
    while True:
        # Prompt the user to input the folder path
        folder_path = input("Enter the full path of the folder containing .doc/.docx files: ").strip()
        
        # Validate folder path
        if not os.path.isdir(folder_path):
            print("Invalid folder path. Please check and try again.")
            continue

        # Convert .doc files to .docx
        for file_name in os.listdir(folder_path):
            if file_name.endswith(".doc") and not file_name.endswith(".docx"):
                file_path = os.path.join(folder_path, file_name)
                print(f"Converting: {file_name}")
                try:
                    convert_doc_to_docx(file_path)
                    print(f"Converted {file_name} to .docx")
                except Exception as e:
                    print(f"Failed to convert {file_name}: {e}")
        
        # Ask if the user wants to use default settings
        use_default_settings = input("Use default settings? [default: start from row 2; take name from col 1, take content from col 3] (y/n): ").strip().lower() in ["y", "yes", "1"]

        if use_default_settings:
            start_row = 1  # Default: Start reading from the second row
            col1_index = 0  # Default: First column
            col3_index = 2  # Default: Third column
        else:
            # Let the user provide custom settings
            start_row = int(input("Enter the row to start reading from (e.g., 1 for the second row): ").strip())
            col1_index = int(input("Enter the column index to extract as Column 1 (e.g., 0 for the first column): ").strip())
            col3_index = int(input("Enter the column index to extract as Column 2 (e.g., 2 for the third column): ").strip())

        # Process each .docx file in the folder
        for file_name in os.listdir(folder_path):
            if file_name.endswith(".docx"):
                file_path = os.path.join(folder_path, file_name)
                print(f"Processing: {file_name}")
                
                # Create a new Excel workbook for this file
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Extracted Data"

                current_row = 1
                
                # Open the Word document
                doc = Document(file_path)
                
                for table in doc.tables:
                    for row in table.rows[start_row:]:
                        # Extract text from user-specified columns
                        col1_text = row.cells[col1_index].text.replace(".", "-") if len(row.cells) > col1_index else ""  # Replace "." with "-"
                        col3_text = row.cells[col3_index].text if len(row.cells) > col3_index else ""
                        
                        # Write to Excel
                        ws.cell(row=current_row, column=1, value=col1_text)
                        ws.cell(row=current_row, column=2, value=col3_text)
                        current_row += 1

                # Save the Excel file with the same name as the Word document
                output_file_name = os.path.splitext(file_name)[0] + ".xlsx"
                output_path = os.path.join(folder_path, output_file_name)
                wb.save(output_path)
                print(f"Data saved to {output_path}")

        print("All files in the folder have been processed successfully.")

        # Ask if the user wants to process another folder
        process_another = input("Do you want to process another folder? (y/n): ").strip().lower()
        if process_another not in ["y", "yes", "1"]:
            print("Exiting the program. Goodbye!")
            break

if __name__ == "__main__":
    process_docx_to_excel()